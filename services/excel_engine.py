"""엑셀 처리 엔진 - Import/Export"""

from datetime import date, datetime
from decimal import Decimal
from typing import List, Optional, Tuple, Dict, Any
from pathlib import Path
import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from sqlmodel import Session, select

from database.models import (
    Contract, Company, MonthlyBilling, CodeMapping
)
from utils.constants import (
    EXCEL_COLUMN_MAPPING, BillingCycle, CompanyType, ContractStatus
)
from utils.parsing_utils import (
    parse_date, parse_amount, parse_boolean,
    parse_warehouse_code, parse_notes_for_rules,
    is_total_row, extract_period_from_item_name
)
from utils.date_utils import parse_billing_timing


class ExcelEngine:
    """엑셀 처리 엔진 - 사내 표준 템플릿 유지"""

    # 템플릿 헤더 (A~S 컬럼)
    TEMPLATE_HEADERS = [
        '창고', '코드', '업체명', '품목명',
        '계약시작일', '계약종료일', '월계약금액', '청구금액',
        '부가세', '합계', '외주업체', '외주금액',
        '이익', '발행시기', '매출일자', '요청일자',
        '매입일자', '특이사항', '자동갱신'
    ]

    def __init__(self, session: Session):
        self.session = session

    def import_from_excel(
        self,
        file_path: str,
        sheet_name: str = "매월 유지보수"
    ) -> Tuple[List[dict], List[dict]]:
        """엑셀에서 데이터 Import

        Args:
            file_path: 엑셀 파일 경로
            sheet_name: 시트명

        Returns:
            (imported_records, errors)
        """
        wb = load_workbook(file_path, data_only=True)

        if sheet_name not in wb.sheetnames:
            # 첫 번째 시트 사용
            ws = wb.active
        else:
            ws = wb[sheet_name]

        records = []
        errors = []

        # 헤더 행 찾기 (첫 번째 행이 헤더라고 가정)
        header_row = 1
        headers = [cell.value for cell in ws[header_row]]

        for row_idx in range(header_row + 1, ws.max_row + 1):
            row_data = {}
            row_errors = []

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                col_letter = get_column_letter(col_idx)

                if col_letter in EXCEL_COLUMN_MAPPING:
                    field_name = EXCEL_COLUMN_MAPPING[col_letter]
                    row_data[field_name] = cell.value

            # 빈 행 스킵
            if not row_data.get('company_code') and not row_data.get('company_name'):
                continue

            # 합계 행 스킵
            if is_total_row(row_data):
                break  # 합계 이후는 메모/지침 블록

            # 데이터 파싱 및 변환
            parsed_record, parse_errors = self._parse_row_data(row_data, row_idx)

            if parse_errors:
                row_errors.extend(parse_errors)

            records.append({
                'row': row_idx,
                'raw_data': row_data,
                'parsed_data': parsed_record,
                'errors': row_errors
            })

            if row_errors:
                errors.extend([{
                    'row': row_idx,
                    'error': e
                } for e in row_errors])

        wb.close()
        return (records, errors)

    def _parse_row_data(
        self,
        row_data: dict,
        row_idx: int
    ) -> Tuple[dict, List[str]]:
        """행 데이터 파싱"""
        parsed = {}
        errors = []

        # 창고 코드
        wh_code, wh_error = parse_warehouse_code(row_data.get('warehouse_code'))
        parsed['warehouse_code'] = wh_code
        if wh_error:
            errors.append(wh_error)

        # 업체 코드/명
        parsed['company_code'] = str(row_data.get('company_code', '')).strip()
        parsed['company_name'] = str(row_data.get('company_name', '')).strip()
        parsed['item_name'] = str(row_data.get('item_name', '')).strip()

        # 계약기간
        start_date, start_error = parse_date(row_data.get('contract_start_date'))
        parsed['contract_start'] = start_date
        if start_error:
            errors.append(start_error)

        end_date, end_error = parse_date(row_data.get('contract_end_date'))
        parsed['contract_end'] = end_date
        if end_error:
            errors.append(end_error)

        # 금액
        monthly_amt, monthly_err, monthly_formula = parse_amount(row_data.get('monthly_amount'))
        parsed['monthly_amount'] = monthly_amt or Decimal("0")
        parsed['monthly_amount_formula'] = monthly_formula
        if monthly_err:
            errors.append(monthly_err)

        billing_amt, billing_err, billing_formula = parse_amount(row_data.get('billing_amount'))
        parsed['billing_amount'] = billing_amt or Decimal("0")
        parsed['billing_amount_formula'] = billing_formula
        if billing_err:
            errors.append(billing_err)

        vat_amt, vat_err, _ = parse_amount(row_data.get('vat_amount'))
        parsed['vat_amount'] = vat_amt or Decimal("0")

        total_amt, total_err, _ = parse_amount(row_data.get('total_amount'))
        parsed['total_amount'] = total_amt or Decimal("0")

        # 외주
        parsed['outsourcing_company'] = str(row_data.get('outsourcing_company', '')).strip()
        outsourcing_amt, outsourcing_err, _ = parse_amount(row_data.get('outsourcing_amount'))
        parsed['outsourcing_amount'] = outsourcing_amt or Decimal("0")
        if outsourcing_err:
            errors.append(outsourcing_err)

        profit_amt, profit_err, _ = parse_amount(row_data.get('profit'))
        parsed['profit'] = profit_amt or Decimal("0")

        # 발행시기
        billing_timing = str(row_data.get('billing_timing', '')).strip()
        parsed['billing_timing'] = billing_timing
        parsed['billing_timing_parsed'] = parse_billing_timing(billing_timing)

        # 일자
        sales_date, _ = parse_date(row_data.get('sales_date'))
        parsed['sales_date'] = sales_date

        request_date, _ = parse_date(row_data.get('request_date'))
        parsed['request_date'] = request_date

        # 매입일자 (다건 허용)
        from utils.parsing_utils import parse_purchase_dates
        purchase_dates, purchase_err = parse_purchase_dates(row_data.get('purchase_date'))
        parsed['purchase_dates'] = purchase_dates
        if purchase_err:
            errors.append(purchase_err)

        # 특이사항
        notes = str(row_data.get('notes', '')).strip()
        parsed['notes'] = notes
        parsed['notes_parsed'] = parse_notes_for_rules(notes)

        # 자동갱신
        parsed['auto_renewal'] = parse_boolean(row_data.get('auto_renewal'))

        # 품목명에서 기간 추출 (다개월 선청구)
        period_months, period_desc = extract_period_from_item_name(parsed['item_name'])
        parsed['cover_months'] = period_months
        parsed['period_description'] = period_desc

        return (parsed, errors)

    def save_imported_data(
        self,
        records: List[dict],
        update_existing: bool = False
    ) -> Tuple[int, int, List[str]]:
        """Import된 데이터 저장

        Returns:
            (created_count, updated_count, errors)
        """
        created = 0
        updated = 0
        errors = []

        for record in records:
            if record.get('errors'):
                continue  # 에러가 있는 행 스킵

            parsed = record['parsed_data']

            try:
                # 업체 조회/생성
                company = self._get_or_create_company(
                    parsed['company_code'],
                    parsed['company_name'],
                    parsed['warehouse_code']
                )

                # 계약 조회/생성
                contract = self._get_or_create_contract(
                    company, parsed, update_existing
                )

                if contract.id:
                    updated += 1
                else:
                    created += 1

            except Exception as e:
                errors.append(f"Row {record['row']}: {str(e)}")

        self.session.commit()
        return (created, updated, errors)

    def _get_or_create_company(
        self,
        code: str,
        name: str,
        warehouse_code: Optional[str]
    ) -> Company:
        """업체 조회 또는 생성"""
        statement = select(Company).where(Company.code == code)
        company = self.session.exec(statement).first()

        if not company:
            company = Company(
                code=code,
                name=name,
                company_type=CompanyType.SALES.value,
                warehouse_code=warehouse_code
            )
            self.session.add(company)
            self.session.flush()

        return company

    def _get_or_create_contract(
        self,
        company: Company,
        parsed: dict,
        update_existing: bool
    ) -> Contract:
        """계약 조회 또는 생성"""
        statement = select(Contract).where(
            Contract.company_id == company.id,
            Contract.item_name == parsed['item_name']
        )
        contract = self.session.exec(statement).first()

        if contract and update_existing:
            # 기존 계약 업데이트
            contract.contract_start = parsed['contract_start']
            contract.contract_end = parsed['contract_end']
            contract.monthly_amount = parsed['monthly_amount']
            contract.billing_timing = parsed['billing_timing']
            contract.billing_timing_parsed = json.dumps(
                parsed['billing_timing_parsed'], ensure_ascii=False
            )
            contract.auto_renewal = parsed['auto_renewal']
            contract.notes = parsed['notes']
            contract.notes_parsed = json.dumps(
                parsed['notes_parsed'], ensure_ascii=False
            )
            contract.updated_at = datetime.now()

        elif not contract:
            # 새 계약 생성
            status = ContractStatus.ACTIVE.value
            if parsed['contract_start'] is None and parsed['contract_end'] is None:
                status = ContractStatus.PERIOD_UNDEFINED.value

            # 역발행 체크
            is_reverse = (
                parsed['billing_timing_parsed'].get('is_reverse_billing', False) or
                parsed['notes_parsed'].get('is_reverse_billing', False)
            )

            contract = Contract(
                company_id=company.id,
                item_name=parsed['item_name'],
                contract_start=parsed['contract_start'],
                contract_end=parsed['contract_end'],
                monthly_amount=parsed['monthly_amount'],
                billing_timing=parsed['billing_timing'],
                billing_timing_parsed=json.dumps(
                    parsed['billing_timing_parsed'], ensure_ascii=False
                ),
                auto_renewal=parsed['auto_renewal'],
                is_reverse_billing=is_reverse,
                default_outsourcing_amount=parsed['outsourcing_amount'],
                status=status,
                notes=parsed['notes'],
                notes_parsed=json.dumps(
                    parsed['notes_parsed'], ensure_ascii=False
                )
            )
            self.session.add(contract)
            self.session.flush()

        return contract

    def export_monthly_billing(
        self,
        year: int,
        month: int,
        file_path: str
    ) -> str:
        """월별 청구 엑셀 Export

        Args:
            year: 청구년도
            month: 청구월
            file_path: 저장 경로

        Returns:
            저장된 파일 경로
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "매월 유지보수"

        # 헤더 스타일
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 헤더 작성
        for col_idx, header in enumerate(self.TEMPLATE_HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 청구 데이터 조회
        statement = select(MonthlyBilling).where(
            MonthlyBilling.billing_year == year,
            MonthlyBilling.billing_month == month
        ).order_by(MonthlyBilling.id)

        billings = self.session.exec(statement).all()

        # 코드 매핑 조회
        code_mappings = {
            cm.code: cm.name
            for cm in self.session.exec(select(CodeMapping)).all()
        }

        # 데이터 작성
        row_idx = 2
        totals = {
            'billing_amount': Decimal("0"),
            'vat_amount': Decimal("0"),
            'total_amount': Decimal("0"),
            'outsourcing_amount': Decimal("0"),
            'profit': Decimal("0")
        }

        for billing in billings:
            contract = billing.contract
            if not contract:
                continue

            company = contract.company

            # 창고 코드
            warehouse_code = company.warehouse_code if company else ""
            ws.cell(row=row_idx, column=1).value = warehouse_code

            # 업체 정보
            ws.cell(row=row_idx, column=2).value = company.code if company else ""
            ws.cell(row=row_idx, column=3).value = company.name if company else ""
            ws.cell(row=row_idx, column=4).value = contract.item_name

            # 계약기간
            ws.cell(row=row_idx, column=5).value = contract.contract_start
            ws.cell(row=row_idx, column=6).value = contract.contract_end

            # 금액
            ws.cell(row=row_idx, column=7).value = float(contract.monthly_amount)
            ws.cell(row=row_idx, column=8).value = float(billing.final_amount)
            ws.cell(row=row_idx, column=9).value = float(billing.vat_amount)
            ws.cell(row=row_idx, column=10).value = float(billing.total_amount)

            # 외주
            ws.cell(row=row_idx, column=11).value = ""  # 외주업체명은 별도 조회 필요
            ws.cell(row=row_idx, column=12).value = float(billing.outsourcing_amount)
            ws.cell(row=row_idx, column=13).value = float(billing.profit)

            # 발행시기
            if contract.is_reverse_billing:
                ws.cell(row=row_idx, column=14).value = "역발행"
            else:
                ws.cell(row=row_idx, column=14).value = contract.billing_timing

            # 일자
            ws.cell(row=row_idx, column=15).value = billing.sales_date
            ws.cell(row=row_idx, column=16).value = billing.request_date
            ws.cell(row=row_idx, column=17).value = None  # 매입일자

            # 특이사항
            ws.cell(row=row_idx, column=18).value = contract.notes

            # 자동갱신
            ws.cell(row=row_idx, column=19).value = "O" if contract.auto_renewal else "X"

            # 합계 누적
            totals['billing_amount'] += billing.final_amount
            totals['vat_amount'] += billing.vat_amount
            totals['total_amount'] += billing.total_amount
            totals['outsourcing_amount'] += billing.outsourcing_amount
            totals['profit'] += billing.profit

            row_idx += 1

        # 합계 행
        ws.cell(row=row_idx, column=3).value = "합계"
        ws.cell(row=row_idx, column=3).font = Font(bold=True)
        ws.cell(row=row_idx, column=8).value = float(totals['billing_amount'])
        ws.cell(row=row_idx, column=9).value = float(totals['vat_amount'])
        ws.cell(row=row_idx, column=10).value = float(totals['total_amount'])
        ws.cell(row=row_idx, column=12).value = float(totals['outsourcing_amount'])
        ws.cell(row=row_idx, column=13).value = float(totals['profit'])

        # 합계 행 스타일
        for col_idx in range(1, 20):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.font = Font(bold=True)

        # 열 너비 조정
        column_widths = [8, 10, 20, 25, 12, 12, 12, 12, 10, 12, 15, 12, 12, 15, 12, 12, 12, 30, 8]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # 저장
        output_path = Path(file_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(file_path)
        wb.close()

        return file_path

    def create_template(self, file_path: str) -> str:
        """빈 템플릿 생성"""
        wb = Workbook()
        ws = wb.active
        ws.title = "매월 유지보수"

        # 헤더 스타일
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')

        for col_idx, header in enumerate(self.TEMPLATE_HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 열 너비
        column_widths = [8, 10, 20, 25, 12, 12, 12, 12, 10, 12, 15, 12, 12, 15, 12, 12, 12, 30, 8]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        output_path = Path(file_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(file_path)
        wb.close()

        return file_path
