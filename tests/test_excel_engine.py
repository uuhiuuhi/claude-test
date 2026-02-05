"""엑셀 엔진 테스트"""

import pytest
from datetime import date
from decimal import Decimal
from pathlib import Path
import tempfile

from services.excel_engine import ExcelEngine
from services.billing_engine import BillingEngine
from database.models import MonthlyBilling
from utils.constants import BillingStatus


class TestExcelExport:
    """엑셀 Export 테스트"""

    def test_export_monthly_billing(self, session, sample_contract):
        """월별 청구 엑셀 Export"""
        # 청구 생성
        billing_engine = BillingEngine(session)
        billings, _ = billing_engine.generate_monthly_billings(2024, 6)
        billing_engine.save_billings(billings)

        # Export
        excel_engine = ExcelEngine(session)

        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = Path(tmpdir) / "test_export.xlsx"
            result_path = excel_engine.export_monthly_billing(2024, 6, str(file_path))

            assert Path(result_path).exists()

            # 파일 검증
            from openpyxl import load_workbook
            wb = load_workbook(result_path)
            ws = wb.active

            # 헤더 확인
            assert ws.cell(row=1, column=1).value == "창고"
            assert ws.cell(row=1, column=3).value == "업체명"
            assert ws.cell(row=1, column=8).value == "청구금액"

            # 데이터 확인
            assert ws.cell(row=2, column=8).value == 1000000.0

            wb.close()

    def test_export_includes_totals(self, session, sample_contract, sample_quarterly_contract):
        """Export에 합계 포함"""
        billing_engine = BillingEngine(session)

        # 월 청구 생성
        billings1, _ = billing_engine.generate_monthly_billings(2024, 6)
        billing_engine.save_billings(billings1)

        # 분기 청구는 6월이 대상이므로 포함됨
        excel_engine = ExcelEngine(session)

        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = Path(tmpdir) / "test_totals.xlsx"
            excel_engine.export_monthly_billing(2024, 6, str(file_path))

            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            ws = wb.active

            # 합계 행 찾기
            last_row = ws.max_row
            assert ws.cell(row=last_row, column=3).value == "합계"

            wb.close()

    def test_create_template(self, session):
        """빈 템플릿 생성"""
        excel_engine = ExcelEngine(session)

        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = Path(tmpdir) / "template.xlsx"
            result_path = excel_engine.create_template(str(file_path))

            assert Path(result_path).exists()

            from openpyxl import load_workbook
            wb = load_workbook(result_path)
            ws = wb.active

            # 헤더만 존재
            assert ws.cell(row=1, column=1).value == "창고"
            assert ws.cell(row=2, column=1).value is None

            wb.close()


class TestExcelImport:
    """엑셀 Import 테스트"""

    def test_import_from_excel(self, session):
        """엑셀 Import"""
        excel_engine = ExcelEngine(session)

        with tempfile.TemporaryDirectory() as tmpdir:
            # 테스트 파일 생성
            file_path = Path(tmpdir) / "test_import.xlsx"

            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "매월 유지보수"

            # 헤더
            headers = ['창고', '코드', '업체명', '품목명', '계약시작일', '계약종료일',
                      '월계약금액', '청구금액', '부가세', '합계', '외주업체', '외주금액',
                      '이익', '발행시기', '매출일자', '요청일자', '매입일자', '특이사항', '자동갱신']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # 데이터
            data = ['105', 'C001', '테스트업체', '유지보수', '2024-01-01', '2024-12-31',
                   1000000, 1000000, 100000, 1100000, '', 0,
                   1000000, '말일', '', '', '', '', 'O']
            for col, value in enumerate(data, 1):
                ws.cell(row=2, column=col, value=value)

            wb.save(file_path)
            wb.close()

            # Import
            records, errors = excel_engine.import_from_excel(str(file_path))

            assert len(records) == 1
            assert records[0]['parsed_data']['company_name'] == '테스트업체'
            assert records[0]['parsed_data']['monthly_amount'] == Decimal("1000000")

    def test_import_stops_at_total_row(self, session):
        """합계 행에서 Import 중단"""
        excel_engine = ExcelEngine(session)

        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = Path(tmpdir) / "test_total.xlsx"

            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active

            # 헤더
            ws.cell(row=1, column=1, value='창고')
            ws.cell(row=1, column=2, value='코드')
            ws.cell(row=1, column=3, value='업체명')

            # 데이터
            ws.cell(row=2, column=2, value='C001')
            ws.cell(row=2, column=3, value='업체1')

            ws.cell(row=3, column=2, value='C002')
            ws.cell(row=3, column=3, value='업체2')

            # 합계 행
            ws.cell(row=4, column=3, value='합계')

            # 합계 아래 메모 (무시되어야 함)
            ws.cell(row=5, column=3, value='메모: 참고사항')

            wb.save(file_path)
            wb.close()

            records, _ = excel_engine.import_from_excel(str(file_path))

            # 합계 전까지만 Import
            assert len(records) == 2


class TestExcelRoundTrip:
    """엑셀 Export/Import 왕복 테스트"""

    def test_export_import_roundtrip(self, session, sample_contract):
        """Export 후 Import 데이터 일치"""
        # 청구 생성 및 저장
        billing_engine = BillingEngine(session)
        billings, _ = billing_engine.generate_monthly_billings(2024, 6)
        billing_engine.save_billings(billings)

        excel_engine = ExcelEngine(session)

        with tempfile.TemporaryDirectory() as tmpdir:
            file_path = Path(tmpdir) / "roundtrip.xlsx"

            # Export
            excel_engine.export_monthly_billing(2024, 6, str(file_path))

            # Import
            records, errors = excel_engine.import_from_excel(str(file_path))

            # 데이터 검증
            assert len(records) == 1
            parsed = records[0]['parsed_data']

            assert parsed['company_name'] == "테스트고객"
            assert parsed['monthly_amount'] == Decimal("1000000")
